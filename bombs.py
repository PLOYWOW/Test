from openpyxl import load_workbook
workbook = load_workbook(filename="Bombss.xlsx")
workbook.sheetnames

sheet = workbook.active

sheet.title

char = ["A","B","C","D","E","F","G","H","I","J","K","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
count = len(char)
#print(count)

#เคลียค่า
for j in range(1,27):
    for r in range(len(char)):
        sheet[char[r]+str(j)] = ""

#สุ่มระเบิด
import random

a = input("Enter Number of Player(s): ")
#print(a)
#print(type(a))

z = int(a) #input
k = 1
for i in range(z):
    x = random.randint(0,24) #A-Z
    y = random.randint(1,26) #1-26
    print(k,char[x],y)
    k = k + 1

    sheet[char[x]+str(y)] = "Bomb"

#รับค่าสุ่ม
ran = input("Input random: ")
#print(ran)
#print(type(ran))

chk = sheet[ran].value
#print(chk)

if chk == "Bomb":
    print("Game Over!!!")

if chk != "Bomb":
    sheet[ran] = " "
    while chk != "Bomb":
        ran = input("Input random: ")
        chk = sheet[ran].value
        #print(chk)
        if chk == " ":
            print("Duplicated")
            continue
        if chk == "Bomb":
            print("Game Over!!!")
            break
        sheet[ran] = " "

workbook.save(filename="Bombss.xlsx")